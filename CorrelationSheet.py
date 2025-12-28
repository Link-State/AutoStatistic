from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.styles import Font as CellFont
from openpyxl.styles import Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties, Font
from Stat import Stat
from Util import *
from FundamentalSheet import FundamentalSheet
import re

class CorrelationSheet() :
    def __init__(self, stat:Stat, category:str):
        self.stat = stat
        self.category = category
        self.work_sheet:Worksheet = self.stat.statExcel.create_sheet(title=self.category)
        self.fundamental_table_coordinate:dict = dict()
        self.correlation_table_coordinate:dict = dict()

        table_title_coordinate = "B2"

        # 전반적인 테이블 생성
        coordinates:dict = FundamentalSheet.createFundamentalTable(self.stat, self.work_sheet, self.category, table_title_coordinate)
        if len(coordinates) <= 0 :
            print("CorrelationSheet.py - 1")
            return
        self.fundamental_table_coordinate = coordinates
        cell:Cell = self.work_sheet[coordinates["sum"][0]]
        table_title_coordinate = cell.offset(self.stat.table_gap + 1, 0).coordinate

        # 상관분석 테이블 생성
        for f_cat in self.stat.fundamental_category :
            coordinates = CorrelationSheet.createCorrelationTable(self.stat, self.work_sheet, self.category, f_cat, table_title_coordinate)
            if len(coordinates) <= 0 :
                print("CorrelationSheet.py - 2")
                return
            self.correlation_table_coordinate[f_cat] = coordinates
            cell:Cell = self.work_sheet[coordinates["sum"][0]]
            table_title_coordinate = cell.offset(self.stat.table_gap + 1, 0).coordinate
        
        # 가로가 가장 긴 테이블을 기준으로 차트 위치 결정
        cell:Cell = self.work_sheet[self.fundamental_table_coordinate["title"][1]]
        maximum_column = cell.column
        for f_cat in self.stat.fundamental_category :
            cell:Cell = self.work_sheet[self.correlation_table_coordinate[f_cat]["title"][1]]
            if maximum_column < cell.column :
                maximum_column = cell.column
        cell:Cell = self.work_sheet.cell(row=2, column=maximum_column+self.stat.table_chart_gap+1)

        # 전반적인 차트
        data_coordinate = self.fundamental_table_coordinate["data"]
        category_coordinate = self.fundamental_table_coordinate["category"]
        chart_coordinate = cell.coordinate
        
        FundamentalSheet.createFundamentalChart(self.stat, self.work_sheet, self.category, data_coordinate, category_coordinate, chart_coordinate)
        cell = cell.offset(18+self.stat.chart_gap, 0)
        chart_coordinate = cell.coordinate

        # 상관분석 차트
        for f_cat in self.stat.fundamental_category :
            data_coordinate = (self.correlation_table_coordinate[f_cat]["h_category"][0], self.correlation_table_coordinate[f_cat]["data"][1])
            category_coordinate = self.correlation_table_coordinate[f_cat]["v_category"]
            CorrelationSheet.createCorrelationChart(self.stat, self.work_sheet, self.category, f_cat, data_coordinate, category_coordinate, chart_coordinate)
            cell = cell.offset(14+self.stat.chart_gap, 0)
            chart_coordinate = cell.coordinate
        

    @staticmethod
    def createCorrelationTable(stat:Stat, work_sheet:Worksheet, h_category:str, v_category:str, table_title_coordinate="B2") -> dict :
        if h_category not in stat.fundamental_category and h_category not in stat.correlation_category :
            return dict()
        if v_category not in stat.fundamental_category and v_category not in stat.correlation_category :
            return dict()
        
        is_number = True
        add_offset = 0

        # 표 제목 입력
        cell:Cell = work_sheet[table_title_coordinate]
        cell.value = f"{v_category}별 {h_category}"
        cell = cell.offset(1, 0)
        
        # 대각선 셀 디자인
        diagonal_border = Border(left=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     right=Side(border_style='double'),
                                     bottom=Side(border_style='double'),
                                     diagonalDown=True,
                                     diagonal=Side(border_style='thin'))
        cell.border = diagonal_border

        # 다중 데이터 처리
        dataframe_exploded = stat.dataframe.copy().dropna(subset=[h_category], how='any', axis=0)
        dataframe_exploded[h_category] = stat.dataframe[h_category].apply(
            lambda x:
                list(
                    set(
                        re.sub(f"{stat.multiSelectSplitSymbol}$", "",
                               re.sub(f"^{stat.multiSelectSplitSymbol}", "",
                                      str(x).replace(f"{stat.nullSymbol}{stat.multiSelectSplitSymbol}", f"미응답{stat.multiSelectSplitSymbol}")
                                            .replace(f"{stat.multiSelectSplitSymbol}{stat.nullSymbol}", f"{stat.multiSelectSplitSymbol}미응답")
                                    )
                            ).split(stat.multiSelectSplitSymbol)
                        )
                    ) if str(x).find(stat.multiSelectSplitSymbol) >= 0 else [x])
        dataframe_exploded = dataframe_exploded.explode(h_category)
        h_axis = dataframe_exploded[h_category].unique()
        v_axis = dataframe_exploded[v_category].unique()
        h_axis.sort(axis=0)
        v_axis.sort(axis=0)

        # 데이터 수 집계
        datas = dataframe_exploded.groupby(v_category)[h_category].value_counts().unstack(fill_value=0)

        # 표 제목 셀 디자인
        title_font = CellFont(name='맑은 고딕', bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        title_border = Border(left=Side(border_style='thin'),
                              top=Side(border_style='thin'),
                              right=Side(border_style='thin'),
                              bottom=Side(border_style='double'))

        # 세로 카테고리 셀 디자인
        v_category_font = CellFont(name='맑은 고딕')
        v_axis_alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        v_category_border = Border(left=Side(border_style='thin'),
                                   top=Side(border_style='thin'),
                                   right=Side(border_style='double'),
                                   bottom=Side(border_style='thin'))
        
        # 가로 카테고리 셀 디자인
        h_category_font = CellFont(name='맑은 고딕')
        h_axis_alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        h_category_border = Border(left=Side(border_style='thin'),
                                   top=Side(border_style='thin'),
                                   right=Side(border_style='thin'),
                                   bottom=Side(border_style='double'))
        
        # 데이터 셀 디자인
        data_font = CellFont(name='맑은 고딕')
        data_border = Border(left=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))
        
        # 합계 셀 디자인
        summation_font = CellFont(name='맑은 고딕', bold=True)
        summation_axis_alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        summation_category_border = Border(left=Side(border_style='thin'),
                                           top=Side(border_style='double'),
                                           right=Side(border_style='double'),
                                           bottom=Side(border_style='thin'))
        summation_border = Border(left=Side(border_style='thin'),
                                  top=Side(border_style='double'),
                                  right=Side(border_style='thin'),
                                  bottom=Side(border_style='thin'))
        
        # 평균점수 셀 디자인
        average_font = CellFont(name='맑은 고딕', bold=True)
        average_axis_alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        average_category_border = Border(left=Side(border_style='double'),
                                         top=Side(border_style='thin'),
                                         right=Side(border_style='thin'),
                                         bottom=Side(border_style='double'))
        average_border = Border(left=Side(border_style='double'),
                                top=Side(border_style='thin'),
                                right=Side(border_style='thin'),
                                bottom=Side(border_style='thin'))
        
        # 가로 카테고리 입력
        h_cell = cell.offset(0, 1)
        for h_ax in h_axis :
            if not isNumber(str(h_ax)) :
                is_number = False
            if isSameInteger(h_ax) :
                h_cell.value = f"{int(h_ax)}{stat.unit[h_category]}"
            else :
                h_cell.value = f"{h_ax}{stat.unit[h_category]}"

            # 가로 카테고리 글꼴, 정렬, 테두리 디자인
            h_cell.font = h_category_font
            h_cell.alignment = h_axis_alignment
            h_cell.border = h_category_border

            h_cell = h_cell.offset(0, 1)
        
        # 평균점수 셀 글꼴, 정렬, 테두리 디자인
        if is_number :
            add_offset = 1
            h_cell.value = f"평균 (단위:{stat.unit[h_category]})"
            h_cell.font = average_font
            h_cell.alignment = average_axis_alignment
            h_cell.border = average_category_border

        # 표 제목 셀 병합
        title_range:str = f"{cell.offset(-1, 0).coordinate}:{cell.offset(-1, len(h_axis)+add_offset).coordinate}"
        work_sheet.merge_cells(title_range)
        range:tuple[tuple[Cell]] = work_sheet[title_range]
        for rows in range :
            for cl in rows :
                cl.font = title_font
                cl.alignment = title_alignment
                cl.border = title_border
        resizeCellWidth(work_sheet, title_range)
        
        # 세로 카테고리 입력
        v_cell = cell.offset(1, 0)
        for v_ax in v_axis :
            if isSameInteger(v_ax) :
                v_cell.value = f"{int(v_ax)}{stat.unit[v_category]}"
            else :
                v_cell.value = f"{v_ax}{stat.unit[v_category]}"
            
            # 세로 카테고리 글꼴, 정렬, 테두리 디자인
            v_cell.font = v_category_font
            v_cell.alignment = v_axis_alignment
            v_cell.border = v_category_border
            
            v_cell = v_cell.offset(1, 0)
        
        # 합계 셀 글꼴, 정렬, 테두리 디자인
        v_cell.value = "합계"
        v_cell.font = summation_font
        v_cell.alignment = summation_axis_alignment
        v_cell.border = summation_category_border
        
        v_cell = cell.offset(1, 1)
        d_cell = cell.offset(1, 1)
        
        # 집계 데이터 입력
        for v_ax in v_axis :
            for h_ax in h_axis :
                d_cell.value = datas.loc[v_ax, h_ax]
                d_cell.font = data_font
                d_cell.border = data_border
                d_cell = d_cell.offset(0, 1)
            if is_number :
                weight = datas.loc[v_ax, :].sort_index(axis=0).values
                weighted_value = h_axis * weight
                d_cell.value = round(weighted_value.sum() / weight.sum(), 1)
                d_cell.font = average_font
                d_cell.border = average_border
            v_cell = v_cell.offset(1, 0)
            d_cell = v_cell.offset(0, 0)
        
        # 합계 입력
        for h_ax in h_axis :
            d_cell.value = datas.loc[:, h_ax].sum()
            d_cell.font = summation_font
            d_cell.border = summation_border
            d_cell = d_cell.offset(0, 1)

        # 좌표 반환
        result = dict()
        cell:Cell = work_sheet[table_title_coordinate]
        result["title"] = [cell.coordinate, cell.offset(0, len(h_axis)+add_offset).coordinate]
        result["diagonal"] = [cell.offset(1, 0).coordinate, cell.offset(1, 0).coordinate]
        result["v_category"] = [cell.offset(2, 0).coordinate, cell.offset(len(v_axis)+1, 0).coordinate]
        result["h_category"] = [cell.offset(1, 1).coordinate, cell.offset(1, len(h_axis)).coordinate]
        result["data"] = [cell.offset(2, 1).coordinate, cell.offset(len(v_axis)+1, len(h_axis)).coordinate]
        result["sum"] = [cell.offset(len(v_axis)+2, 0).coordinate, cell.offset(len(v_axis)+2, len(h_axis)).coordinate]
        result["average"] = [cell.offset(1, len(h_axis)+add_offset).coordinate, cell.offset(len(v_axis)+1, len(h_axis)+add_offset).coordinate]
        result["entire"] = [cell.coordinate, cell.offset(len(v_axis)+2, len(h_axis)+add_offset).coordinate]

        return result
    
    @staticmethod
    def createCorrelationChart(stat:Stat, work_sheet:Worksheet, h_category:str, v_category:str, data_coordinate:tuple[str], category_coordinate:tuple[str], chart_coordinate="B2") :
        
        # chart_space = ChartSpace(roundedCorners=False)
        # line_properties = LineProperties(round=False)
        # graphical_properties = GraphicalProperties(ln=line_properties)

        # 차트 제목 글꼴 설정
        font = Font(typeface="맑은 고딕")
        character_properties = CharacterProperties(latin=font, sz=1800, b=True)
        paragraph_properties = ParagraphProperties(defRPr=character_properties)

        # 차트 준비
        chart = BarChart()
        chart.roundedCorners = False
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.title = f"{v_category}별 {h_category}"
        chart.title.tx.rich.p[0].pPr = paragraph_properties
        chart.title.overlay = False
        chart.type = "col"
        chart.grouping = "percentStacked"
        chart.overlap = 100
        chart.legend.overlay = False
        
        # 참조데이터 준비
        data_start_cell:Cell = work_sheet[data_coordinate[0]]
        data_end_cell:Cell = work_sheet[data_coordinate[1]]
        category_start_cell:Cell = work_sheet[category_coordinate[0]]
        category_end_cell:Cell = work_sheet[category_coordinate[1]]
        
        data = Reference(work_sheet,
                         min_col=data_start_cell.column,
                         min_row=data_start_cell.row,
                         max_col=data_end_cell.column,
                         max_row=data_end_cell.row)
        cats = Reference(work_sheet,
                         min_col=category_start_cell.column,
                         min_row=category_start_cell.row,
                         max_row=category_end_cell.row)

        # 차트 데이터 및 카테고리 추가
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 12.8524
        chart.height = 7.7978

        # 열이 1개일 경우, 숨겨진 더미데이터 추가
        if len(chart.series) == 1 :
            dummydata = Reference(work_sheet, min_col=data_start_cell.column+1, min_row=data_start_cell.row, max_col=data_end_cell.column+1, max_row=data_end_cell.row)
            chart.add_data(dummydata, titles_from_data=True)
            chart.series[1].spPr.noFill = True

        work_sheet.add_chart(chart, chart_coordinate)

        return