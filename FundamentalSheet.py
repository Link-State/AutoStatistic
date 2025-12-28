from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.styles import Font as CellFont
from openpyxl.styles import Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
from Stat import Stat
from Util import *
import re

class FundamentalSheet() :
    def __init__(self, stat:Stat):
        self.stat = stat
        self.work_sheet:Worksheet = self.stat.statExcel.create_sheet(title="기본통계")
        self.table_coordinate:dict = dict()

        # 표 만들기
        table_title_coordinate = "B2"
        for cat in self.stat.fundamental_category :
            coordinates:dict = FundamentalSheet.createFundamentalTable(self.stat, self.work_sheet, cat, table_title_coordinate)
            if len(coordinates) <= 0 :
                print("FundamentalSheet.py - 1")
                return
            self.table_coordinate[cat] = coordinates
            cell:Cell = self.work_sheet[coordinates["sum"][0]]
            table_title_coordinate = cell.offset(self.stat.table_gap + 1, 0).coordinate

        # 차트 만들기
        chart_coordinate:str = self.table_coordinate[self.stat.fundamental_category[0]]["title"][1]
        cell:Cell = self.work_sheet[chart_coordinate]
        cell:Cell = self.work_sheet.cell(row=cell.row, column=cell.column+self.stat.table_chart_gap+1)
        chart_coordinate = cell.coordinate

        for cat in self.stat.fundamental_category :
            if cat not in self.table_coordinate :
                continue
            data_coordinate = self.table_coordinate[cat]["data"]
            category_coordinate = self.table_coordinate[cat]["category"]
            FundamentalSheet.createFundamentalChart(self.stat, self.work_sheet, cat, data_coordinate, category_coordinate, chart_coordinate)
            cell = cell.offset(18 + self.stat.chart_gap, 0)
            chart_coordinate = cell.coordinate
    
    @staticmethod
    def createFundamentalTable(stat:Stat, work_sheet:Worksheet, category:str, table_title_coordinate="B2") -> dict :
        if category not in stat.fundamental_category and category not in stat.correlation_category :
            return dict()
        
        # 표 제목
        cell:Cell = work_sheet[table_title_coordinate]
        cell.value = f"전반적인 {category}"
        cell = cell.offset(1, 0)

        # 표 제목 셀 병합
        title_range:str = f"{cell.offset(-1, 0).coordinate}:{cell.offset(-1, 1).coordinate}"
        work_sheet.merge_cells(title_range)
        range:tuple[tuple[Cell]] = work_sheet[title_range]

        # 표 제목 셀 디자인
        title_font = CellFont(name='맑은 고딕', bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        title_border = Border(left=Side(border_style='thin'),
                              top=Side(border_style='thin'),
                              right=Side(border_style='thin'),
                              bottom=Side(border_style='double'))
        for rows in range :
            for cl in rows :
                cl.font = title_font
                cl.alignment = title_alignment
                cl.border = title_border
        resizeCellWidth(work_sheet, title_range)
        
        # 다중 데이터 처리
        dataframe_exploded = stat.dataframe.copy()
        dataframe_exploded[category] = stat.dataframe[category].apply(
            lambda x:
                list(
                    set(
                        re.sub(f"{stat.multiSelectSplitSymbol}$", "",
                            re.sub(f"^{stat.multiSelectSplitSymbol}", "",
                                    str(x).replace(f"{stat.nullSymbol}{stat.multiSelectSplitSymbol}", f"미응답{stat.multiSelectSplitSymbol}")
                                          .replace(f"{stat.multiSelectSplitSymbol}{stat.nullSymbol}", f"{stat.multiSelectSplitSymbol}미응답")
                                )
                            ).split(stat.multiSelectSplitSymbol)
                        )
                    ) if str(x).find(stat.multiSelectSplitSymbol) >= 0 else [x])
        dataframe_exploded = dataframe_exploded.explode(category)
        axis = dataframe_exploded[category].unique()
        axis.sort(axis=0)

        # 데이터 수 집계
        datas = dataframe_exploded[category].value_counts()

        # 카테고리 셀 디자인
        category_font = CellFont(name='맑은 고딕')
        axis_alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        category_border = Border(left=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             right=Side(border_style='double'),
                             bottom=Side(border_style='thin'))
        
        # 합계 셀 디자인
        summation_font = CellFont(name='맑은 고딕', bold=True)
        summation_category_border = Border(left=Side(border_style='thin'),
                             top=Side(border_style='double'),
                             right=Side(border_style='double'),
                             bottom=Side(border_style='thin'))
        summation_border = Border(left=Side(border_style='thin'),
                             top=Side(border_style='double'),
                             right=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))
        
        # 데이터 셀 디자인
        data_font = CellFont(name='맑은 고딕')
        data_border = Border(left=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))
        
        # 데이터 입력
        for ax in axis :
            if isSameInteger(ax) :
                cell.value = f"{int(ax)}{stat.unit[category]}"
            else :
                cell.value = f"{ax}{stat.unit[category]}"
            cell.offset(0, 1).value = datas[ax]

            # 각 카테고리 셀에 스타일 적용
            cell.font = category_font
            cell.alignment = axis_alignment
            cell.border = category_border

            # 각 데이터 셀에 스타일 적용
            cell.offset(0, 1).font = data_font
            cell.offset(0, 1).border = data_border

            cell = cell.offset(1, 0)
        
        cell.value = "합계"
        cell.offset(0, 1).value = datas.sum()
        
        # 합계 셀 스타일 적용
        cell.font = summation_font
        cell.alignment = axis_alignment
        cell.border = summation_category_border
        cell.offset(0, 1).font = summation_font
        cell.offset(0, 1).border = summation_border

        # 좌표 반환
        result = dict()
        cell:Cell = work_sheet[table_title_coordinate]
        result["title"] = [cell.coordinate, cell.offset(0, 1).coordinate]
        result["category"] = [cell.offset(1, 0).coordinate, cell.offset(len(axis), 0).coordinate]
        result["data"] = [cell.offset(1, 1).coordinate, cell.offset(len(axis), 1).coordinate]
        result["sum"] = [cell.offset(len(axis)+1, 0).coordinate, cell.offset(len(axis)+1, 1).coordinate]
        result["entire"] = [cell.coordinate, cell.offset(len(axis)+1, 1).coordinate]

        return result
    
    @staticmethod
    def createFundamentalChart(stat:Stat, work_sheet:Worksheet, category:str, data_coordinate:tuple[str], category_coordinate:tuple[str], chart_coordinate="B2") :
        
        # 차트 제목 글꼴 설정
        title_font = Font(typeface="맑은 고딕")
        title_character_properties = CharacterProperties(latin=title_font, sz=1800, b=True)
        title_paragraph_properties = ParagraphProperties(defRPr=title_character_properties)

        # 차트 준비
        chart = PieChart()
        chart.roundedCorners = False
        chart.title = f"전반적인 {category}"
        chart.title.tx.rich.p[0].pPr = title_paragraph_properties
        chart.title.overlay = False
        chart.style = 2
        chart.legend = None

        # 참조데이터 준비
        data_start_cell:Cell = work_sheet[data_coordinate[0]]
        data_end_cell:Cell = work_sheet[data_coordinate[1]]
        category_start_cell:Cell = work_sheet[category_coordinate[0]]
        category_end_cell:Cell = work_sheet[category_coordinate[1]]
        
        data = Reference(work_sheet,
                         min_col=data_start_cell.column,
                         min_row=data_start_cell.row-1,
                         max_col=data_end_cell.column,
                         max_row=data_end_cell.row)
        
        cats = Reference(work_sheet,
                         min_col=category_start_cell.column,
                         min_row=category_start_cell.row,
                         max_row=category_end_cell.row)

        # 차트 데이터 및 카테고리 추가
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 12.8524
        chart.height = 10.1854

        # 차트 데이터라벨 글꼴 설정
        fontsize = 700
        fontbold = False
        data_amount = data_end_cell.row - data_start_cell.row
        if data_amount <= 1 :
            fontsize = 1400
            fontbold = True
        elif data_amount <= 3 :
            fontsize = 1000
            fontbold = True
        elif data_amount <= 5 :
            fontsize = 900
        elif data_amount <= 7 :
            fontsize = 800

        labels_font = Font(typeface="Calibri")
        labels_character_properties = CharacterProperties(latin=labels_font, sz=fontsize, b=fontbold)
        labels_paragraph_properties = ParagraphProperties(defRPr=labels_character_properties)
        labels_paragraph = Paragraph(pPr=labels_paragraph_properties)
        labels_paragraphs = [labels_paragraph]
        labels_rich = RichText(p=labels_paragraphs)

        datalabel_list = DataLabelList(delete=False)
        datalabel_list.showSerName = False
        datalabel_list.showCatName = True
        datalabel_list.showVal = False
        datalabel_list.showPercent = True
        datalabel_list.showLegendKey = False
        datalabel_list.separator = "\n"
        datalabel_list.dLblPos = 'bestFit'
        datalabel_list.txPr = labels_rich
        chart.dataLabels = datalabel_list

        work_sheet.add_chart(chart, chart_coordinate)

        return