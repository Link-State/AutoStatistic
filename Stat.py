import sys
import os
import openpyxl as Excel
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from Exception import StatisticException


class Stat() :
    def __init__(self, refer:str, name="무제1", tableGap=2, chartGap=2, nullSymbol="-", catStartPoint="B3", dataStartPoint="B4", multiSelectSplitSymbol=";") :
        from Category import Category

        self.statFilePath: str = None
        self.statFileName: str = name
        self.statExcel: Workbook = None
        self.referenceFile: str = refer
        self.moveDir: tuple[int, int] = (0, 1)
        self.table_gap = tableGap
        self.chart_gap = chartGap
        self.nullSymbol = nullSymbol
        self.catStartPoint: str = catStartPoint
        self.dataStartPoint: str = dataStartPoint
        self.multiSelectSplitSymbol: str = multiSelectSplitSymbol
        self.sheets: dict = dict()
        self.categories: dict = dict()
        self.comparisonCategories: list = list()
        self.generalCategories: list = list()

        if not os.path.exists(self.referenceFile) :
            detail = self.referenceFile
            self.referenceFile = None
            raise StatisticException(3, detail)
        
        self.referenceExcel: Workbook = Excel.load_workbook(refer)

        # 항목행 불러와서 같이 분석할 항목 선택받기
        inquirySheetName: str = self.referenceExcel.sheetnames[0]
        inquirySheet: Worksheet = self.referenceExcel[inquirySheetName]
        categoryCell: Cell = inquirySheet[catStartPoint]
        
        while True :
            if type(categoryCell.value) != type(str()) :
                break
            
            catName = str(categoryCell.value).strip()
            if catName == "" :
                break
            
            if catName != "" :
                category = Category(stat=self, cell=categoryCell)
                self.categories[category.name] = category
                categoryCell = categoryCell.offset(self.moveDir[0], self.moveDir[1])
        
        return
    
    # 카테고리 분리
    def divideCategory(self) :
        from Category import Category

        self.comparisonCategories.clear()
        self.generalCategories.clear()

        keys = self.categories
        for key in keys :
            cat:Category = self.categories[key]
            if cat.isComparison :
                self.comparisonCategories.append(cat)
            else :
                self.generalCategories.append(cat)
        
        return
    
    # 시트 생성
    def createSheet(self) :
        from Sheet import FundamentalSheet, CorrelationSheet
        
        sheetNames = list()
        self.sheets.clear()

        # 기본통계 시트 생성
        sheet = FundamentalSheet(stat=self, comparisonCats=self.comparisonCategories)
        self.sheets[sheet.name] = sheet
        sheetNames.append(sheet.name)

        # 상관분석 시트 생성
        for cat in self.generalCategories :
            sheet = CorrelationSheet(stat=self, targetCat=cat, comparisonCats=self.comparisonCategories)
            self.sheets[sheet.name] = sheet
            sheetNames.append(sheet.name)

        return sheetNames
    
    def inflate(self) :
        from Sheet import Sheet, ChartOrganizeSheet

        self.statExcel = Workbook()
        ws = self.statExcel.active
        
        # 엑셀파일 내 시트 생성
        sheetNames:list = self.createSheet()
        for sheetName in self.sheets.keys() :
            sheet: Sheet = self.sheets[sheetName]
            sheet.inflate()

        # 차트 총괄 시트 생성
        sheet = ChartOrganizeSheet(stat=self, sheetNames=sheetNames)
        sheet.inflate()

        return
    
    def show(self) :
        self.inflate()
        
        try :
            self.statExcel.save(self.statFilePath)
        except PermissionError :
            raise StatisticException(0, self.statFilePath)
        finally :
            self.statExcel.close()
        
        # 생성 완료!

        return
